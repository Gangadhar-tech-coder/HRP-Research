import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.spatial.distance import pdist, squareform
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ============================================
# CONFIGURATION
# ============================================
DATASET_PATH = "C:\\Users\\anju&janu\\OneDrive\\kavya\\HRP Research\\Research\\calculations_5f_748.csv"

# NEW: Ask user for number of sampleys to use
print("\n--- Sample Selection ---")
use_subset = input("Do you want to use a subset of samples? (y/n): ").strip().lower()

REDUCTION_PERCENTAGES = [70, 50, 30, 10]
ALPHA_VALUES = np.arange(0.0, 1.1, 0.1)
RANDOM_STATE = 42

# ============================================
# CREATE OUTPUT FOLDER STRUCTURE
# ============================================
# Extract dataset name from path
dataset_name = os.path.splitext(os.path.basename(DATASET_PATH))[0]
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
output_folder = f"results_{dataset_name}_{timestamp}"

# Create folder structure
os.makedirs(output_folder, exist_ok=True)
images_folder = os.path.join(output_folder, "images")
os.makedirs(images_folder, exist_ok=True)

print(f"\nOutput folder created: {output_folder}")
print(f"Images will be saved to: {images_folder}")

# ============================================
# LOAD AND PREPARE DATASET
# ============================================
print(f"\nLoading dataset from: {DATASET_PATH}")
df = pd.read_csv(DATASET_PATH)

# Automatically detect and remove non-numeric columns
print("\nDetecting column types...")
non_numeric_cols = df.select_dtypes(exclude=[np.number]).columns.tolist()
if non_numeric_cols:
    print(f"Found non-numeric columns: {non_numeric_cols}")
    print(f"Automatically removing: {', '.join(non_numeric_cols)}")
    df = df.select_dtypes(include=[np.number])
else:
    print("All columns are numeric.")

# NEW: Sample selection logic
print(f"\nFull dataset shape: {df.shape}")
print(f"Total samples available: {df.shape[0]}, Features: {df.shape[1]}")

if use_subset == 'y':
    n_samples_to_use = int(input(f"Enter number of samples to use (max {df.shape[0]}): ").strip())
    if n_samples_to_use > df.shape[0]:
        print(f"Warning: Requested {n_samples_to_use} samples but only {df.shape[0]} available. Using all samples.")
        n_samples_to_use = df.shape[0]
    
    # Randomly sample n_samples_to_use rows
    np.random.seed(RANDOM_STATE)
    sampled_indices = np.random.choice(df.shape[0], size=n_samples_to_use, replace=False)
    df = df.iloc[sampled_indices].reset_index(drop=True)
    print(f"\nUsing {n_samples_to_use} randomly sampled instances")
else:
    print("\nUsing all available samples")

X = df.to_numpy()

print(f"\nFinal dataset shape: {X.shape}")
print(f"Samples: {X.shape[0]}, Features: {X.shape[1]}\n")

n_samples, n_features = X.shape

# ============================================
# RANDOM PROJECTION IMPLEMENTATIONS
# ============================================
def create_nrp_matrix(n_features, n_components, random_state):
    np.random.seed(random_state)
    return np.random.randn(n_features, n_components) / np.sqrt(n_components)

def create_pmrp_matrix(n_features, n_components, random_state):
    np.random.seed(random_state)
    return np.random.choice([-1, 1], size=(n_features, n_components)) / np.sqrt(n_components)

def create_hrp_matrix(n_features, n_components, alpha, random_state):
    R1 = create_nrp_matrix(n_features, n_components, random_state)
    R2 = create_pmrp_matrix(n_features, n_components, random_state + 1)
    return alpha * R1 + (1 - alpha) * R2

def compute_distortion(X_original, X_projected, n_samples_dist=5000):
    """Compute distortion using sampled distances to reduce memory usage."""
    n_samples = X_original.shape[0]
    ratios = []
    
    # Sample random pairs of points
    n_pairs = min(n_samples_dist, n_samples * (n_samples - 1) // 2)
    
    np.random.seed(RANDOM_STATE)  # For reproducibility
    for _ in range(n_pairs):
        i, j = np.random.choice(n_samples, 2, replace=False)
        
        d_orig = np.linalg.norm(X_original[i] - X_original[j])
        d_proj = np.linalg.norm(X_projected[i] - X_projected[j])
        
        if d_orig > 1e-10:
            ratio = np.abs(d_proj / d_orig - 1)
            ratios.append(ratio)
    
    return np.mean(ratios) if ratios else 0.0

# ============================================
# EXCEL FORMATTING FUNCTION
# ============================================
def format_excel_sheet(worksheet, has_header=True):
    """Apply professional formatting to Excel sheet"""
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    border_thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Format header row
    if has_header:
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
    
    # Format data rows
    for row in worksheet.iter_rows(min_row=2 if has_header else 1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
            if isinstance(cell.value, float):
                cell.number_format = '0.00000000'
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

# ============================================
# EXCEL WRITER SETUP
# ============================================
excel_path = os.path.join(output_folder, f"{dataset_name}_results.xlsx")
writer = pd.ExcelWriter(excel_path, engine='openpyxl')

# Create dataset info sheet
dataset_info_data = [
    ['Dataset Name', dataset_name],
    ['Dataset Path', DATASET_PATH],
    ['Total Available Samples', df.shape[0]],
    ['Total Available Features', df.shape[1]],
    ['Samples Used', X.shape[0]],
    ['Features Used', X.shape[1]],
    ['Random State', RANDOM_STATE],
    ['Timestamp', timestamp]
]
dataset_info = pd.DataFrame(dataset_info_data, columns=['Property', 'Value'])
dataset_info.to_excel(writer, sheet_name='Dataset Info', index=False)

# ============================================
# MAIN EXPERIMENT LOOP
# ============================================
best_alphas = {}
best_distortions = {}
all_results = []  # Store all detailed results

for reduction_percentage in REDUCTION_PERCENTAGES:
    k = int(n_features * reduction_percentage / 100)
    
    if k < 1:
        print(f"Skipping reduction {reduction_percentage}% (k={k} too small)")
        continue
    
    # Print terminal header
    print(f"\n{'='*80}")
    print(f"Reduction: {reduction_percentage}% | Reduced Dimension: {k}")
    print(f"{'='*80}")
    print(f"{'Alpha':<10} {'NRP Distortion':<20} {'PMRP Distortion':<20} {'HRP Distortion':<20}")
    print(f"{'-'*80}")
    
    # OPTIMIZED: Calculate NRP and PMRP only ONCE per reduction
    R_nrp = create_nrp_matrix(n_features, k, RANDOM_STATE)
    X_nrp = X @ R_nrp
    distortion_nrp = compute_distortion(X, X_nrp)
    
    R_pmrp = create_pmrp_matrix(n_features, k, RANDOM_STATE)
    X_pmrp = X @ R_pmrp
    distortion_pmrp = compute_distortion(X, X_pmrp)
    
    # Store distortions for plotting
    nrp_distortions = []
    pmrp_distortions = []
    hrp_distortions = []
    
    # Data for Excel sheet
    reduction_data = []
    
    # OPTIMIZED: Only calculate HRP for each alpha
    for alpha in ALPHA_VALUES:
        # NRP and PMRP distortions are constant across all alphas
        nrp_distortions.append(distortion_nrp)
        pmrp_distortions.append(distortion_pmrp)
        
        # HRP - Calculate for each alpha value
        R_hrp = create_hrp_matrix(n_features, k, alpha, RANDOM_STATE)
        X_hrp = X @ R_hrp
        distortion_hrp = compute_distortion(X, X_hrp)
        hrp_distortions.append(distortion_hrp)
        
        # Print to terminal
        print(f"{alpha:<10.1f} {distortion_nrp:<20.10f} {distortion_pmrp:<20.10f} {distortion_hrp:<20.10f}")
        
        # Store for Excel
        reduction_data.append({
            'Alpha': alpha,
            'NRP Distortion': distortion_nrp,
            'PMRP Distortion': distortion_pmrp,
            'HRP Distortion': distortion_hrp
        })
        
        # Store for summary
        all_results.append({
            'Reduction %': reduction_percentage,
            'k': k,
            'Alpha': alpha,
            'NRP Distortion': distortion_nrp,
            'PMRP Distortion': distortion_pmrp,
            'HRP Distortion': distortion_hrp
        })
    
    # Save this reduction's data to Excel
    df_reduction = pd.DataFrame(reduction_data)
    sheet_name = f'Reduction_{reduction_percentage}pct'
    df_reduction.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Find best alpha for HRP
    min_idx = np.argmin(hrp_distortions)
    best_alpha = ALPHA_VALUES[min_idx]
    min_distortion = hrp_distortions[min_idx]
    
    best_alphas[reduction_percentage] = best_alpha
    best_distortions[reduction_percentage] = min_distortion
    
    # Print terminal footer
    print(f"\nBest Alpha for {reduction_percentage}% Reduction: {best_alpha:.1f}")
    print(f"Minimum HRP Distortion: {min_distortion:.10f}\n")
    
    # Plot distortion vs alpha
    plt.figure(figsize=(12, 7))
    plt.plot(ALPHA_VALUES, nrp_distortions, 'o-', label='NRP', linewidth=2.5, markersize=10)
    plt.plot(ALPHA_VALUES, pmrp_distortions, 's-', label='PMRP', linewidth=2.5, markersize=10)
    plt.plot(ALPHA_VALUES, hrp_distortions, '^-', label='HRP', linewidth=2.5, markersize=10)
    plt.axvline(best_alpha, color='red', linestyle='--', linewidth=2.5, alpha=0.7, label=f'Best alpha={best_alpha:.1f}')
    
    plt.xlabel('Alpha', fontsize=16, fontweight='bold')
    plt.ylabel('Distance Distortion', fontsize=16, fontweight='bold')
    plt.title(f'{dataset_name} - Reduction {reduction_percentage}% (k={k})', 
              fontsize=18, fontweight='bold', pad=20)
    plt.legend(fontsize=14, loc='best')
    plt.grid(True, alpha=0.3, linestyle='--')
    plt.tight_layout()
    
    # Save to images folder
    image_path = os.path.join(images_folder, f'distortion_vs_alpha_{reduction_percentage}pct.png')
    plt.savefig(image_path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"Saved graph: {image_path}")

# ============================================
# SAVE SUMMARY SHEET TO EXCEL
# ============================================
summary_data = []
for reduction_percentage in REDUCTION_PERCENTAGES:
    if reduction_percentage in best_alphas:
        k = int(n_features * reduction_percentage / 100)
        summary_data.append({
            'Reduction %': reduction_percentage,
            'Target Dimension (k)': k,
            'Best Alpha': best_alphas[reduction_percentage],
            'Minimum HRP Distortion': best_distortions[reduction_percentage]
        })

df_summary = pd.DataFrame(summary_data)
df_summary.to_excel(writer, sheet_name='Summary', index=False)

# Save all detailed results
df_all_results = pd.DataFrame(all_results)
df_all_results.to_excel(writer, sheet_name='All Results', index=False)

# Close and reopen to apply formatting
writer.close()

# ============================================
# APPLY EXCEL FORMATTING
# ============================================
print("\nApplying Excel formatting...")
workbook = load_workbook(excel_path)

# Format each sheet
for sheet_name in workbook.sheetnames:
    worksheet = workbook[sheet_name]
    format_excel_sheet(worksheet, has_header=True)

workbook.save(excel_path)
print(f"Excel file saved with formatting: {excel_path}")

# ============================================
# FINAL SUMMARY TABLE (TERMINAL)
# ============================================
print(f"\n{'='*80}")
print("FINAL SUMMARY")
print(f"{'='*80}")
print(f"{'Reduction %':<15} {'k':<15} {'Best Alpha':<15} {'Minimum Distortion':<25}")
print(f"{'-'*80}")

for reduction_percentage in REDUCTION_PERCENTAGES:
    if reduction_percentage in best_alphas:
        k = int(n_features * reduction_percentage / 100)
        print(f"{reduction_percentage:<15} {k:<15} {best_alphas[reduction_percentage]:<15.1f} {best_distortions[reduction_percentage]:<25.10f}")

print(f"{'='*80}")

# ============================================
# FINAL SUMMARY BAR GRAPH
# ============================================
reductions = list(best_alphas.keys())
alphas = list(best_alphas.values())

plt.figure(figsize=(12, 7))
bars = plt.bar(reductions, alphas, color='steelblue', edgecolor='black', linewidth=2, alpha=0.8, width=8)

for i, (reduction, alpha) in enumerate(zip(reductions, alphas)):
    plt.text(reduction, alpha + 0.03, f'{alpha:.1f}', ha='center', va='bottom', 
             fontsize=14, fontweight='bold')

plt.xlabel('Dimensionality Reduction (%)', fontsize=16, fontweight='bold')
plt.ylabel('Best Alpha', fontsize=16, fontweight='bold')
plt.title(f'{dataset_name} - Best Alpha vs Dimensionality Reduction', 
          fontsize=18, fontweight='bold', pad=20)
plt.ylim(0, 1.2)
plt.grid(True, alpha=0.3, axis='y', linestyle='--')
plt.tight_layout()

# Save summary bar graph
summary_image_path = os.path.join(images_folder, 'best_alpha_vs_reduction.png')
plt.savefig(summary_image_path, dpi=300, bbox_inches='tight')
plt.close()
print(f"Saved summary graph: {summary_image_path}")

# ============================================
# CREATE README FILE
# ============================================
readme_path = os.path.join(output_folder, "README.txt")
with open(readme_path, 'w', encoding='utf-8') as f:
    f.write("="*80 + "\n")
    f.write("HYBRID RANDOM PROJECTION EXPERIMENT RESULTS\n")
    f.write("="*80 + "\n\n")
    f.write(f"Dataset: {dataset_name}\n")
    f.write(f"Dataset Path: {DATASET_PATH}\n")
    f.write(f"Timestamp: {timestamp}\n")
    f.write(f"Original Dimensions: {n_samples} samples x {n_features} features\n\n")
    f.write("="*80 + "\n")
    f.write("FILES IN THIS FOLDER:\n")
    f.write("="*80 + "\n")
    f.write(f"1. {dataset_name}_results.xlsx - Complete results in Excel format\n")
    f.write(f"2. images/ - All generated graphs\n")
    f.write(f"3. README.txt - This file\n\n")
    f.write("="*80 + "\n")
    f.write("EXCEL SHEETS:\n")
    f.write("="*80 + "\n")
    f.write("- Dataset Info: Dataset metadata and configuration\n")
    f.write("- Summary: Best alpha for each reduction percentage\n")
    f.write("- All Results: Complete data for all experiments\n")
    for reduction_percentage in REDUCTION_PERCENTAGES:
        if reduction_percentage in best_alphas:
            f.write(f"- Reduction_{reduction_percentage}pct: Detailed alpha vs distortion table\n")
    f.write("\n")
    f.write("="*80 + "\n")
    f.write("BEST RESULTS:\n")
    f.write("="*80 + "\n")
    f.write(f"{'Reduction %':<15} {'k':<10} {'Best Alpha':<15} {'Min Distortion':<20}\n")
    f.write("-"*80 + "\n")
    for reduction_percentage in REDUCTION_PERCENTAGES:
        if reduction_percentage in best_alphas:
            k = int(n_features * reduction_percentage / 100)
            f.write(f"{reduction_percentage:<15} {k:<10} {best_alphas[reduction_percentage]:<15.1f} {best_distortions[reduction_percentage]:<20.10f}\n")

print(f"Created README: {readme_path}")

print("\n" + "="*80)
print("EXPERIMENT COMPLETED SUCCESSFULLY!")
print("="*80)
print(f"\nAll results saved to: {output_folder}/")
print(f"- Excel file: {dataset_name}_results.xlsx (with professional formatting)")
print(f"- Images: images/ folder ({len(best_alphas) + 1} graphs)")
print(f"- README: README.txt")